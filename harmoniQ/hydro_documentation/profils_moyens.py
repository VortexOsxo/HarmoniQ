# -*- coding: utf-8 -*-
"""
Génération des profils journaliers moyens depuis les données CEHQ.
Produit :
  - 6 fichiers Excel de profils pour barrages réservoirs (format laforge_débit.xlsx)
  - 11 fichiers Excel pour barrages fil de l'eau
"""
import pandas as pd
import numpy as np
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font

def lire_debits_cehq(chemin):
    encodages = ["windows-1252", "latin-1", "utf-8"]
    for enc in encodages:
        try:
            with open(chemin, "r", encoding=enc) as f:
                lignes = f.readlines()
            break
        except: continue
    header_idx = None
    for i, l in enumerate(lignes):
        if "Station" in l and "Date" in l:
            header_idx = i
            break
    data_str = "".join(lignes[header_idx:])
    df = pd.read_csv(StringIO(data_str), sep=r"\s+", engine="python", on_bad_lines="skip")
    mots_cles = ["debit", "débit", "m3", "m³"]
    colonne_debit = None
    for col in df.columns:
        col_norm = col.lower().replace("(","").replace(")","").replace("³","3")
        if any(m in col_norm for m in mots_cles):
            colonne_debit = col
            break
    df["Date"] = pd.to_datetime(df[df.columns[1]], errors="coerce")
    debits = pd.to_numeric(df[colonne_debit], errors="coerce")
    result = pd.DataFrame({"Date": df["Date"], "Debit": debits})
    return result.dropna().reset_index(drop=True)

# ─── FICHIERS SOURCE ────────────────────────────────────────────────────────
fichiers_reservoirs = {
    "Caniapiscau":    "CANIAPISCAUUUU.txt",
    "Petite_Baleine": "APPROXPETITEBALEINE.txt",
    "A_la_Baleine":   "ALABALEINEEEE.txt",
    "Petit_Mecatina": "PETITMECATINAAAA.txt",
    "Magpie":         "MAGPIEEEE.txt",
    "Grande_Baleine": "GRANDEBALEINEEEEE.txt",
}

fichiers_fil_eau = {
    "Caniapisca-3": "Caniapiscau",
    "Caniapisca-4": "Caniapiscau",
    "Caniapisca-5": "Caniapiscau",
    "Caniapisca-6": "Caniapiscau",
    "Caniapisca-7": "Caniapiscau",
    "AlaBaleine-1": "A_la_Baleine",
    "AlaBaleine-2": "A_la_Baleine",
    "AlaBaleine-3": "A_la_Baleine",
    "GdeBaleine-3": "Grande_Baleine",
    "PtMecatina-3": "Petit_Mecatina",
    "PtMecatina-4": "Petit_Mecatina",
}

dates_2025 = pd.date_range("2025-01-01", "2025-12-31", freq="D")

def construire_profil(chemin):
    df = lire_debits_cehq(chemin)
    df = df[df["Debit"] > 0].copy()
    df["jour"] = df["Date"].dt.dayofyear
    return df.groupby("jour")["Debit"].mean().reindex(range(1, 366), fill_value=None).interpolate()

def sauvegarder_excel(nom_fichier, profil):
    wb = Workbook()
    ws = wb.active
    font = Font(name="Arial", size=10)
    for i, date in enumerate(dates_2025):
        debit = round(profil.get(date.dayofyear, profil.iloc[-1]), 1)
        ws.cell(row=i+1, column=1, value=date.to_pydatetime()).font = font
        ws.cell(row=i+1, column=1).number_format = "YYYY/MM/DD"
        ws.cell(row=i+1, column=2, value=debit).font = font
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    wb.save(nom_fichier)

if __name__ == "__main__":
    # Adapter les chemins selon votre dossier
    DATA_DIR = "."

    profils = {}
    for nom, fichier in fichiers_reservoirs.items():
        profils[nom] = construire_profil(f"{DATA_DIR}/{fichier}")
        sauvegarder_excel(f"{nom}_debit.xlsx", profils[nom])
        print(f"Créé : {nom}_debit.xlsx")

    for nom_barrage, riviere in fichiers_fil_eau.items():
        sauvegarder_excel(f"{nom_barrage}.xlsx", profils[riviere])
        print(f"Créé : {nom_barrage}.xlsx")
